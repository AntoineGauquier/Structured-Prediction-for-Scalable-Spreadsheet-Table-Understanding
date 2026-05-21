"""XLSX file loader."""

import os
import tempfile
import shutil
import zipfile
import openpyxl
import pandas as pd
from .base import SpreadsheetLoader
from ..utils.compression import is_empty_value
from pathlib import Path

_FULL_ROW_COL_THRESHOLD = 1000


def _compute_effective_max_col(sheet):
    """
    Compute the effective last column with content, mirroring the LibreOffice
    macro ComputeUsedRangeWithMergedCells.
    """
    raw_max_col = sheet.max_column or 0
    raw_max_row = sheet.max_row or 0

    # Fast heuristic: column count is already reasonable, no correction needed.
    if raw_max_col < _FULL_ROW_COL_THRESHOLD:
        return raw_max_col

    full_row_merge_rows = set() # 1-based row numbers covered by full-row merges
    non_frm_max_col = 0 # max col from non-full-row merges

    try:
        for mr in sheet.merged_cells.ranges:
            # openpyxl uses 1-based column indices.
            if mr.min_col == 1 and mr.max_col >= _FULL_ROW_COL_THRESHOLD:
                for r in range(mr.min_row, mr.max_row + 1):
                    full_row_merge_rows.add(r)
            else:
                non_frm_max_col = max(non_frm_max_col, mr.max_col)
    except Exception:
        # merged_cells unavailable (read_only=True fallback): can't correct.
        return raw_max_col

    if not full_row_merge_rows:
        # max_column was large but no actual full-row merges found: trust it.
        return raw_max_col

    # Cap column scan at _FULL_ROW_COL_THRESHOLD to match LO's behaviour andavoid iterating all 16384 xlsx columns.
    scan_col_limit = min(raw_max_col, _FULL_ROW_COL_THRESHOLD)

    for row_tuple in sheet.iter_rows(
        max_row=raw_max_row,
        max_col=scan_col_limit,
        values_only=False,
    ):
        if not row_tuple:
            continue
        if row_tuple[0].row in full_row_merge_rows:
            continue
        for cell in row_tuple:
            if cell.value is not None:
                non_frm_max_col = max(non_frm_max_col, cell.column)

    return non_frm_max_col if non_frm_max_col > 0 else 1


def _load_workbook_stripped_styles(xlsx_path):
    """
    Fallback loader for XLSX files whose xl/styles.xml is corrupted or
    incompatible with openpyxl.  Replaces the styles with a minimal valid
    stylesheet so the rest of the file can be read.

    Mirrors the same function as in annotation_processing.py so both
    pipelines handle the same files.

    Returns (workbook, tmp_path, read_only_flag).
    The caller must delete tmp_path after use.
    """

    minimal_styles = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<fonts count="1"><font/></fonts>'
        b'<fills count="2">'
        b'<fill><patternFill patternType="none"/></fill>'
        b'<fill><patternFill patternType="gray125"/></fill>'
        b'</fills>'
        b'<borders count="1"><border/></borders>'
        b'<cellStyleXfs count="1">'
        b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>'
        b'</cellStyleXfs>'
        b'<cellXfs count="1">'
        b'<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>'
        b'</cellXfs>'
        b'</styleSheet>'
    )

    tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(tmp_fd)

    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin, zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = (minimal_styles
                        if item.filename == 'xl/styles.xml'
                        else zin.read(item.filename))
                zout.writestr(item, data)
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=False)
            return wb, tmp_path, False
        except Exception:
            wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
            return wb, tmp_path, True
    except Exception:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


def build_merged_cell_map_xlsx(sheet):
    """
    Build {(row_0based, col_0based): (anchor_row_0based, anchor_col_0based)}
    for every cell in every merged region (origin AND covered cells).
    Origin cells map to themselves.

    Requires read_only=False.  Returns {} on read_only=True.
    """
    merged_map = {}
    try:
        for mr in sheet.merged_cells.ranges:
            anchor_row = mr.min_row - 1
            anchor_col = mr.min_col - 1
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    merged_map[(row - 1, col - 1)] = (anchor_row, anchor_col)
    except Exception:
        pass
    return merged_map


def get_xlsx_cell_value(sheet, row, col):
    """
    Return the computed value of an XLSX cell (workbook must use data_only=True).

    Returns None for:
      - Real empty cells
      - Uncalculated formula cells (no cached result)
      - Error cells (data_type == 'e')

    Returns the raw Python value (int, float, str, bool, datetime) otherwise.
    """
    cell = sheet.cell(row + 1, col + 1)   # openpyxl is 1-based

    # Error cells: annotation tool checks data_type == 'e' and returns ''.
    if cell.data_type == 'e':
        return None

    return cell.value # None for empty/uncalculated formula cells


class XLSXLoader(SpreadsheetLoader):
    """Loader for XLSX files."""

    MAX_CONSECUTIVE_EMPTY_ROWS = 200

    def load(self, path, sheet_name=None):
        """
        Load an XLSX file.
        """
        path_obj = Path(path)

        # Ensure .xlsx extension (openpyxl rejects other suffixes).
        if path_obj.suffix.lower() != '.xlsx':
            tmp_fd, tmp_xlsx_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(tmp_fd)
            shutil.copy(path_obj, tmp_xlsx_path)
            load_path = Path(tmp_xlsx_path)
        else:
            load_path = path_obj
            tmp_xlsx_path = None

        wb = None
        read_only = False
        tmp_stripped = None

        try:
            try:
                wb = openpyxl.load_workbook(
                    load_path, data_only=True, read_only=False
                )
            except Exception:
                try:
                    wb = openpyxl.load_workbook(
                        load_path, data_only=True, read_only=True
                    )
                    read_only = True
                except Exception:
                    wb, tmp_stripped, read_only = _load_workbook_stripped_styles(
                        str(load_path)
                    )

            sheet_idx = self._resolve_sheet_index(wb.worksheets, sheet_name)
            sheet = wb.worksheets[sheet_idx]

            # Build merged cell map (requires read_only=False); on read_only fallback the map is empty.
            merged_map = {} if read_only else build_merged_cell_map_xlsx(sheet)

            if read_only:
                effective_max_col = sheet.max_column or 0
            else:
                effective_max_col = _compute_effective_max_col(sheet)

            max_row = sheet.max_row or 0
            cell_map = {}
            data = []
            consecutive_empty = 0

            for row_idx in range(max_row):
                row_data = []
                has_content = False

                for col_idx in range(effective_max_col):
                    if (row_idx, col_idx) in merged_map:
                        anchor_row, anchor_col = merged_map[(row_idx, col_idx)]
                        cell_value = get_xlsx_cell_value(sheet, anchor_row, anchor_col)
                    else:
                        cell_value = get_xlsx_cell_value(sheet, row_idx, col_idx)

                    if not is_empty_value(cell_value):
                        has_content = True

                    row_data.append(cell_value)
                    cell_map[(row_idx, col_idx)] = (row_idx, col_idx)

                if not has_content:
                    consecutive_empty += 1
                    if consecutive_empty >= self.MAX_CONSECUTIVE_EMPTY_ROWS:
                        break
                else:
                    consecutive_empty = 0

                data.append(row_data)

            df = pd.DataFrame(data)
            return df, sheet, wb, merged_map, cell_map

        finally:
            if tmp_xlsx_path and str(load_path) != str(path_obj):
                try:
                    os.unlink(tmp_xlsx_path)
                except OSError:
                    pass
            if tmp_stripped:
                try:
                    os.unlink(tmp_stripped)
                except OSError:
                    pass