from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, List, Optional


# ─────────────────────────────────────────────────────────────────────────────
#  Data container
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class SpreadsheetData:
    values: List[List[Any]]
    number_formats: List[List[str]]
    n_rows: int
    n_cols: int
    sheet_name: Optional[str] = None
    file_path: Optional[str] = None
    merged_cells: List[str] = field(default_factory=list)
    format_attrs: List[List[List[str]]] = field(default_factory=list)


# ─────────────────────────────────────────────────────────────────────────────
#  Magic-byte detection helpers
# ─────────────────────────────────────────────────────────────────────────────

_OLE2_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'
_ZIP_MAGIC  = b'PK\x03\x04'


def _sniff_bytes(path: str) -> bytes:
    with open(path, 'rb') as f:
        return f.read(8)


def _is_ole2(path: str) -> bool:
    return _sniff_bytes(path).startswith(_OLE2_MAGIC)


def _is_zip(path: str) -> bool:
    return _sniff_bytes(path).startswith(_ZIP_MAGIC)


# ─────────────────────────────────────────────────────────────────────────────
#  Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def load_spreadsheet(
    path: str,
    sheet_name: Optional[str] = None,
    mime_type: Optional[str] = None,
) -> SpreadsheetData:
    """Load a spreadsheet from *path*.

    Dispatch order:
      1. MIME type (when provided by the caller)
      2. File extension
      3. Magic bytes

    sheet_name=None → active / first sheet.
    """
    ext = os.path.splitext(path)[1].lower()

    # ── XLSX family ───────────────────────────────────────────────────────
    xlsx_mimes = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel.sheet.macroEnabled.12',
    }
    if (mime_type in xlsx_mimes
            or ext in ('.xlsx', '.xlsm', '.xltx', '.xltm')
            or (ext not in ('.xls', '.ods', '.csv', '.tsv') and _is_zip(path))):
        return _load_xlsx(path, sheet_name)

    # ── XLS (OLE2) ────────────────────────────────────────────────────────
    if (mime_type == 'application/vnd.ms-excel'
            or ext == '.xls'
            or _is_ole2(path)):
        return _load_xls(path, sheet_name)

    # ── ODS ───────────────────────────────────────────────────────────────
    if (mime_type == 'application/vnd.oasis.opendocument.spreadsheet'
            or ext == '.ods'):
        return _load_ods(path, sheet_name)

    # ── CSV / TSV (and text/csv look-alikes) ──────────────────────────────
    if (mime_type in ('text/csv', 'text/x-csv', 'text/tab-separated-values')
            or ext in ('.csv', '.tsv', '.asp', '.aspx')):
        delimiter = '\t' if ext == '.tsv' else ','
        return _load_csv(path, delimiter=delimiter)

    # ── Last resort: try by extension then magic ──────────────────────────
    if ext in ('.xlsx', '.xlsm'):
        return _load_xlsx(path, sheet_name)
    if ext == '.xls':
        return _load_xls(path, sheet_name)
    if ext == '.ods':
        return _load_ods(path, sheet_name)
    if ext in ('.csv', '.tsv'):
        return _load_csv(path)

    raise ValueError(
        f"Cannot determine format for {path!r}. "
        "Pass mime_type= or use a recognised extension."
    )


# ─────────────────────────────────────────────────────────────────────────────
#  XLSX loader (openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def _load_xlsx(path: str, sheet_name: Optional[str]) -> SpreadsheetData:
    from openpyxl import load_workbook

    # Some servers serve xlsx with non-.xlsx extensions; load via BytesIO to
    # bypass openpyxl's extension check.
    with open(path, 'rb') as f:
        raw = f.read()
    wb = load_workbook(BytesIO(raw), data_only=True, read_only=False)

    if sheet_name is None:
        ws = wb.active
    elif sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        raise KeyError(
            f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )

    n_rows = ws.max_row or 0
    n_cols = ws.max_column or 0

    values  = [[None]      * n_cols for _ in range(n_rows)]
    nfs     = [['General'] * n_cols for _ in range(n_rows)]
    fmt_attrs = [[[]        for _ in range(n_cols)] for _ in range(n_rows)]

    for row in ws.iter_rows(min_row=1, max_row=n_rows,
                            min_col=1, max_col=n_cols):
        for cell in row:
            r = cell.row - 1
            c = cell.column - 1
            if not (0 <= r < n_rows and 0 <= c < n_cols):
                continue
            values[r][c] = cell.value
            nfs[r][c] = cell.number_format or 'General'
            fmt_attrs[r][c] = _xlsx_cell_attrs(cell)

    merged = [str(rng) for rng in (ws.merged_cells.ranges if ws.merged_cells else [])]

    return SpreadsheetData(
        values=values, number_formats=nfs,
        n_rows=n_rows, n_cols=n_cols,
        sheet_name=ws.title, file_path=path,
        merged_cells=merged, format_attrs=fmt_attrs,
    )


def _xlsx_cell_attrs(cell) -> List[str]:
    """Extract formatting attributes in the Appendix-F style."""
    attrs: List[str] = []
    try:
        if cell.font and cell.font.bold:
            attrs.append('FontBold')
        if cell.border:
            b = cell.border
            if b.top    and b.top.border_style:    attrs.append('TopBorder')
            if b.bottom and b.bottom.border_style: attrs.append('BottomBorder')
            if b.left   and b.left.border_style:   attrs.append('LeftBorder')
            if b.right  and b.right.border_style:  attrs.append('RightBorder')
        if cell.fill and cell.fill.fill_type not in (None, 'none', 'solid'):
            attrs.append('FillColor')
        elif cell.fill and cell.fill.fill_type == 'solid':
            fg = cell.fill.fgColor
            if fg and fg.type != 'none' and fg.rgb not in ('00000000', 'FFFFFFFF', None):
                attrs.append('FillColor')
    except Exception:
        pass
    return attrs


# ─────────────────────────────────────────────────────────────────────────────
#  XLS loader (xlrd)
# ─────────────────────────────────────────────────────────────────────────────

def _load_xls(path: str, sheet_name: Optional[str]) -> SpreadsheetData:
    import xlrd
    import datetime

    wb = xlrd.open_workbook(path, formatting_info=True)

    if sheet_name is None:
        ws = wb.sheets()[0]
    else:
        names = [s.name for s in wb.sheets()]
        if sheet_name not in names:
            raise KeyError(
                f"Sheet {sheet_name!r} not found. Available: {names}"
            )
        ws = wb.sheet_by_name(sheet_name)

    n_rows, n_cols = ws.nrows, ws.ncols
    values    = [[None]      * n_cols for _ in range(n_rows)]
    nfs       = [['General'] * n_cols for _ in range(n_rows)]
    fmt_attrs = [[[] for _ in range(n_cols)] for _ in range(n_rows)]

    for r in range(n_rows):
        for c in range(n_cols):
            cell = ws.cell(r, c)
            ctype = cell.ctype
            cval  = cell.value

            if ctype == xlrd.XL_CELL_EMPTY or ctype == xlrd.XL_CELL_BLANK:
                values[r][c] = None
            elif ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(cval, wb.datemode)
                    values[r][c] = dt
                except Exception:
                    values[r][c] = cval
            elif ctype == xlrd.XL_CELL_BOOLEAN:
                values[r][c] = bool(cval)
            elif ctype == xlrd.XL_CELL_ERROR:
                values[r][c] = None
            else:
                values[r][c] = cval

            # Number format string
            try:
                xf_idx = ws.cell_xf_index(r, c)
                xf = wb.xf_list[xf_idx]
                fmt = wb.format_map.get(xf.format_key)
                if fmt and fmt.format_str:
                    nfs[r][c] = fmt.format_str
            except Exception:
                pass

    return SpreadsheetData(
        values=values, number_formats=nfs,
        n_rows=n_rows, n_cols=n_cols,
        sheet_name=ws.name, file_path=path,
        merged_cells=[], format_attrs=fmt_attrs,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  ODS loader (odfpy)
# ─────────────────────────────────────────────────────────────────────────────

_ODF_TABLE_NS  = 'urn:oasis:names:tc:opendocument:xmlns:table:1.0'
_ODF_OFFICE_NS = 'urn:oasis:names:tc:opendocument:xmlns:office:1.0'
_ODF_TEXT_NS   = 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'

def _load_ods(path: str, sheet_name: Optional[str]) -> SpreadsheetData:
    from odf.opendocument import load as odf_load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = odf_load(path)
    sheets = doc.spreadsheet.getElementsByType(Table)
    if not sheets:
        raise ValueError(f"No sheets found in ODS file: {path!r}")

    if sheet_name is None:
        ws = sheets[0]
    else:
        matches = [s for s in sheets if s.getAttribute('name') == sheet_name]
        if not matches:
            available = [s.getAttribute('name') for s in sheets]
            raise KeyError(
                f"Sheet {sheet_name!r} not found. Available: {available}"
            )
        ws = matches[0]

    rows_data: List[List[Any]] = []

    for row_elem in ws.getElementsByType(TableRow):
        row_repeat = int(
            row_elem.attributes.get((_ODF_TABLE_NS, 'number-rows-repeated'), 1)
        )

        cell_vals: List[Any] = []
        for cell_elem in row_elem.getElementsByType(TableCell):
            col_repeat = int(
                cell_elem.attributes.get(
                    (_ODF_TABLE_NS, 'number-columns-repeated'), 1)
            )
            val = _ods_cell_value(cell_elem)
            cell_vals.extend([val] * col_repeat)

        for _ in range(row_repeat):
            rows_data.append(list(cell_vals))

    if not rows_data:
        return SpreadsheetData(
            values=[], number_formats=[],
            n_rows=0, n_cols=0,
            sheet_name=ws.getAttribute('name'), file_path=path,
        )

    # Trim trailing all-None rows
    while rows_data and all(v is None for v in rows_data[-1]):
        rows_data.pop()

    n_rows = len(rows_data)
    n_cols = max((len(r) for r in rows_data), default=0)

    # Pad rows to uniform width and trim trailing None columns
    padded = [r + [None] * (n_cols - len(r)) for r in rows_data]

    # Trim trailing empty columns
    while n_cols > 0 and all(padded[r][n_cols - 1] is None for r in range(n_rows)):
        n_cols -= 1
    padded = [r[:n_cols] for r in padded]

    values    = padded
    nfs       = [['General'] * n_cols for _ in range(n_rows)]
    fmt_attrs = [[[] for _ in range(n_cols)] for _ in range(n_rows)]

    return SpreadsheetData(
        values=values, number_formats=nfs,
        n_rows=n_rows, n_cols=n_cols,
        sheet_name=ws.getAttribute('name'), file_path=path,
        merged_cells=[], format_attrs=fmt_attrs,
    )


def _ods_cell_value(cell_elem) -> Any:
    """Extract the Python value from an ODF TableCell element."""
    import datetime

    attrs = cell_elem.attributes
    vtype = attrs.get((_ODF_OFFICE_NS, 'value-type'))

    if vtype == 'float':
        raw = attrs.get((_ODF_OFFICE_NS, 'value'))
        if raw is not None:
            try:
                f = float(raw)
                return int(f) if f == int(f) else f
            except ValueError:
                pass
    elif vtype == 'percentage':
        raw = attrs.get((_ODF_OFFICE_NS, 'value'))
        if raw is not None:
            try:
                return float(raw)
            except ValueError:
                pass
    elif vtype == 'currency':
        raw = attrs.get((_ODF_OFFICE_NS, 'value'))
        if raw is not None:
            try:
                return float(raw)
            except ValueError:
                pass
    elif vtype == 'date':
        raw = attrs.get((_ODF_OFFICE_NS, 'date-value'))
        if raw:
            try:
                return datetime.date.fromisoformat(raw[:10])
            except ValueError:
                pass
    elif vtype == 'time':
        raw = attrs.get((_ODF_OFFICE_NS, 'time-value'))
        if raw:
            return raw  # ISO 8601 duration string
    elif vtype == 'boolean':
        raw = attrs.get((_ODF_OFFICE_NS, 'boolean-value'))
        return raw == 'true' if raw is not None else None
    elif vtype == 'string':
        from odf.text import P
        texts = [str(p) for p in cell_elem.getElementsByType(P)]
        return ''.join(texts) or None

    # Fallback: read display text
    from odf.text import P
    texts = [str(p) for p in cell_elem.getElementsByType(P)]
    text = ''.join(texts).strip()
    return text if text else None


# ─────────────────────────────────────────────────────────────────────────────
#  CSV / TSV loader
# ─────────────────────────────────────────────────────────────────────────────

_ENCODINGS_TO_TRY = ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252', 'utf-16')


def _detect_csv_encoding(path: str) -> str:
    """Try common encodings; return the first one that decodes the whole file."""
    with open(path, 'rb') as f:
        raw = f.read()
    # BOM checks first
    if raw.startswith(b'\xff\xfe') or raw.startswith(b'\xfe\xff'):
        return 'utf-16'
    if raw.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    for enc in _ENCODINGS_TO_TRY:
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return 'latin-1'  # last resort — never raises


def _detect_csv_delimiter(path: str, encoding: str) -> str:
    """Sniff delimiter from first few lines; fall back to comma."""
    with open(path, encoding=encoding, errors='replace', newline='') as f:
        sample = f.read(4096)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        return dialect.delimiter
    except csv.Error:
        return ','


def _load_csv(path: str, delimiter: Optional[str] = None) -> SpreadsheetData:
    encoding  = _detect_csv_encoding(path)
    if delimiter is None:
        delimiter = _detect_csv_delimiter(path, encoding)

    rows: List[List[Any]] = []
    with open(path, encoding=encoding, errors='replace', newline='') as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row in reader:
            rows.append(list(row))

    n_rows = len(rows)
    n_cols = max((len(r) for r in rows), default=0)

    values    = [[None]      * n_cols for _ in range(n_rows)]
    nfs       = [['General'] * n_cols for _ in range(n_rows)]
    fmt_attrs = [[[] for _ in range(n_cols)] for _ in range(n_rows)]

    for r, row in enumerate(rows):
        for c, val in enumerate(row):
            values[r][c] = val if val != '' else None

    return SpreadsheetData(
        values=values, number_formats=nfs,
        n_rows=n_rows, n_cols=n_cols,
        sheet_name=None, file_path=path,
        merged_cells=[], format_attrs=fmt_attrs,
    )


# ─────────────────────────────────────────────────────────────────────────────
#  Convenience constructor
# ─────────────────────────────────────────────────────────────────────────────

def from_matrix(
    values: List[List[Any]],
    number_formats: Optional[List[List[str]]] = None,
    sheet_name: Optional[str] = None,
) -> SpreadsheetData:
    """Build a SpreadsheetData from a plain 2-D list (for tests / programmatic use)."""
    n_rows = len(values)
    n_cols = max((len(r) for r in values), default=0)
    padded = [list(r) + [None] * (n_cols - len(r)) for r in values]
    if number_formats is None:
        nfs = [['General'] * n_cols for _ in range(n_rows)]
    else:
        nfs = [list(r) + ['General'] * (n_cols - len(r)) for r in number_formats]
    fmt_attrs = [[[] for _ in range(n_cols)] for _ in range(n_rows)]
    return SpreadsheetData(
        values=padded, number_formats=nfs,
        n_rows=n_rows, n_cols=n_cols,
        sheet_name=sheet_name, file_path=None,
        merged_cells=[], format_attrs=fmt_attrs,
    )
